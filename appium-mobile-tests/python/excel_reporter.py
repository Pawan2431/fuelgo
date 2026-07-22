import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class PythonExcelReporter:
    def __init__(self, report_dir=None):
        if report_dir is None:
            report_dir = os.path.join(os.path.dirname(__file__), "..", "reports")
        self.report_dir = report_dir
        os.makedirs(self.report_dir, exist_ok=True)

    def generate_report(self, test_results, device_info=None):
        if device_info is None:
            device_info = {
                "deviceName": "Android Emulator (UiAutomator2)",
                "platformName": "Android",
                "platformVersion": "13.0",
                "app": "FuelGo Mobile App"
            }

        wb = Workbook()
        
        # ----------------------------------------------------
        # TAB 1: SUMMARY DASHBOARD
        # ----------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        # Header Title
        ws_summary.merge_cells('A1:E2')
        title_cell = ws_summary['A1']
        title_cell.value = "📱 FuelGo Android Appium E2E Automation Report (Python)"
        title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="0B1A3B", end_color="0B1A3B", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        total_tests = len(test_results)
        passed_tests = sum(1 for t in test_results if t.get('status') == 'PASS')
        failed_tests = sum(1 for t in test_results if t.get('status') == 'FAIL')
        pass_rate = round((passed_tests / total_tests) * 100, 1) if total_tests > 0 else 0
        total_duration = round(sum(t.get('duration', 0) for t in test_results) / 1000, 2)

        ws_summary.append([])
        ws_summary.append(["Execution Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_summary.append(["Target Device", device_info.get("deviceName")])
        ws_summary.append(["Platform", f"{device_info.get('platformName')} {device_info.get('platformVersion')}"])
        ws_summary.append(["App / Package", device_info.get("app")])
        ws_summary.append([])

        ws_summary.append(["KPI Metric", "Result Value", "Benchmark Target"])
        kpi_header_row = ws_summary[9]
        for cell in kpi_header_row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1565D8", end_color="1565D8", fill_type="solid")

        kpis = [
            ["Total Test Cases", total_tests, "Target: 100% Executed"],
            ["Passed Test Cases", passed_tests, "Target: All Pass"],
            ["Failed Test Cases", failed_tests, "0 Failures"],
            ["Pass Rate (%)", f"{pass_rate}%", "Threshold: >= 90%"],
            ["Total Duration (sec)", f"{total_duration}s", "Performance Nominal"]
        ]

        for row in kpis:
            ws_summary.append(row)

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 30
        ws_summary.column_dimensions['C'].width = 25

        # ----------------------------------------------------
        # TAB 2: TEST DETAILS
        # ----------------------------------------------------
        ws_details = wb.create_sheet(title="Detailed Test Logs")
        ws_details.views.sheetView[0].showGridLines = True

        headers = ["#", "Suite", "Test Case Description", "Screen", "Duration (ms)", "Status", "Error Details"]
        ws_details.append(headers)

        header_row = ws_details[1]
        for cell in header_row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for idx, t in enumerate(test_results, start=1):
            row_data = [
                idx,
                t.get('suite', 'E2E Suite'),
                t.get('title'),
                t.get('screen', 'General'),
                t.get('duration', 0),
                t.get('status'),
                t.get('error', 'N/A')
            ]
            ws_details.append(row_data)
            
            # Color status cell
            status_cell = ws_details.cell(row=idx+1, column=6)
            if t.get('status') == 'PASS':
                status_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                status_cell.font = Font(color="2E7D32", bold=True)
            else:
                status_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                status_cell.font = Font(color="C62828", bold=True)

        col_widths = {'A': 6, 'B': 20, 'C': 42, 'D': 18, 'E': 15, 'F': 12, 'G': 40}
        for col, width in col_widths.items():
            ws_details.column_dimensions[col].width = width

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f"Python_Appium_E2E_Report_{timestamp}.xlsx"
        file_path = os.path.join(self.report_dir, file_name)

        wb.save(file_path)
        print(f"\n✅ Python Excel Analysis Report generated successfully:")
        print(f"📂 Path: {file_path}")
        return file_path
